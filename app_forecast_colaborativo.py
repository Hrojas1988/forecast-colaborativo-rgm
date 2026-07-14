"""
Aplicativo de Forecast Colaborativo — Ventas y Unidades por Cliente-SKU
========================================================================
Corre localmente con:
    pip install -r requirements.txt
    streamlit run app_forecast_colaborativo.py

Flujo:
1. Sube tu histórico (Cliente, SKU, Mes, Unidades, Precio + N columnas macro que tú definas).
2. La app ajusta una Regresión Lineal por cada combinación Cliente-SKU:
   Unidades ~ tendencia + estacionalidad (sin/cos) + tus variables macro.
3. Editas en pantalla el escenario macro futuro, el % colaborativo por
   Cliente-SKU-Mes y el precio — todo con inputs interactivos.
4. Descargas el resultado en un Excel con varias hojas.
"""

import io
from datetime import datetime

import numpy as np
import pandas as pd
import altair as alt
import streamlit as st
from sklearn.linear_model import LinearRegression
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

try:
    from streamlit_gsheets import GSheetsConnection
    from gspread.exceptions import WorksheetNotFound
    GSHEETS_DISPONIBLE = True
except ImportError:
    GSHEETS_DISPONIBLE = False

st.set_page_config(page_title="Modelo de Proyección de Presupuesto Colaborativo CBM - CAM SUR", layout="wide")

# -------------------------------------------------------------------
# 0. Utilidades
# -------------------------------------------------------------------
REQUIRED_COLS = ["Cliente", "SKU", "Mes", "Unidades", "Precio"]
ATRIBUTOS_CLIENTE = ["Canal", "Grupo_Cliente", "Pais"]  # opcionales, atributos por Cliente (características, no predictoras)
ATRIBUTOS_SKU_BB = ["Marca", "Categoria"]  # opcionales, atributos por SKU usados en building blocks
ATRIBUTOS_SKU_ID = ["SKU_ID"]  # opcional, identificador de SKU separado del nombre/descripción — solo referencia
ATRIBUTOS_SKU = ATRIBUTOS_SKU_BB + ATRIBUTOS_SKU_ID
ATRIBUTOS_BB = ATRIBUTOS_CLIENTE + ATRIBUTOS_SKU_BB  # dimensiones que sí se grafican en building blocks


def hay_credenciales_gsheets() -> bool:
    try:
        return "connections" in st.secrets and "gsheets" in st.secrets["connections"]
    except Exception:
        return False


def leer_tabla_compartida(conn, hoja: str, default_df: pd.DataFrame, claves: list, nombre_legible: str = None) -> pd.DataFrame:
    """Lee una pestaña del Google Sheet compartido. Si no existe, está vacía, o sus claves
    (Cliente/SKU/Mes/etc.) no corresponden al archivo actualmente cargado — por ejemplo, quedó
    guardada de una corrida anterior con otro archivo — usa el default fresco en su lugar,
    para no arrastrar NaN por combinaciones que ya no existen."""
    try:
        df = conn.read(worksheet=hoja, ttl=60)
        df = df.dropna(how="all")
        if df.empty or not all(c in df.columns for c in claves):
            return default_df

        claves_guardadas = set(map(tuple, df[claves].astype(str).values.tolist()))
        claves_esperadas = set(map(tuple, default_df[claves].astype(str).values.tolist()))
        interseccion = claves_guardadas & claves_esperadas
        # Ojo: el denominador es lo GUARDADO, no la grilla completa esperada — así los guardados
        # parciales (solo las filas que alguien editó, no las ~decenas de miles en cero) validan bien.
        cobertura = len(interseccion) / len(claves_guardadas) if claves_guardadas else 1.0
        if claves_guardadas and cobertura < 0.5:
            if nombre_legible:
                st.sidebar.warning(f"⚠️ Lo guardado antes en '{nombre_legible}' no corresponde a este archivo "
                                    f"(parece de otra corrida u otro histórico) — usando valores nuevos por defecto.")
            return default_df

        # Alinea tipos de columnas numéricas con el default para que el data_editor no truene
        for col in default_df.columns:
            if col in df.columns and pd.api.types.is_numeric_dtype(default_df[col]):
                df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
        faltantes_col = [c for c in default_df.columns if c not in df.columns]
        for c in faltantes_col:
            df[c] = default_df[c].iloc[0] if len(default_df) else ""

        # Aunque haya suficiente cobertura, puede haber filas nuevas (SKUs/meses nuevos) sin guardar
        # todavía — las completamos con el default en vez de dejarlas en NaN.
        df = default_df[claves].merge(df, on=claves, how="left")
        for col in default_df.columns:
            if col not in claves:
                df[col] = df[col].fillna(default_df[col])
        return df[default_df.columns]
    except Exception:
        return default_df


def guardar_tabla_compartida(conn, hoja: str, df: pd.DataFrame):
    """Actualiza la pestaña si ya existe; si no existe todavía en el Sheet, la crea."""
    try:
        conn.update(worksheet=hoja, data=df)
    except WorksheetNotFound:
        conn.create(worksheet=hoja, data=df)


def datos_demo():
    """Genera un histórico ficticio de 24 meses para probar la app sin subir archivo."""
    np.random.seed(7)
    clientes_info = [
        ("Supermercado Central", "Moderno", "Cadenas Nacionales", "Costa Rica"),
        ("Tienda Express", "Tradicional", "Mayoristas", "Panamá"),
    ]
    skus_info = [
        ("SKU-101", "Bebida 600ml", "Marca Azul", "Bebidas"),
        ("SKU-205", "Snack 150g", "Marca Roja", "Snacks"),
    ]
    meses = pd.date_range("2024-07-01", periods=24, freq="MS")
    filas = []
    for i, (cli, canal, grupo, pais) in enumerate(clientes_info):
        for j, (sku_id, sku_nombre, marca, categoria) in enumerate(skus_info):
            base, trend = 900 + i * 300, 4 + j * 2
            for k, mes in enumerate(meses):
                inflacion = 0.006 + 0.002 * np.sin(k / 6) + np.random.normal(0, 0.0008)
                tc = 0.004 + 0.003 * np.sin(k / 9) + np.random.normal(0, 0.001)
                season = 50 * np.sin(2 * np.pi * mes.month / 12)
                unidades = max(base + trend * k + season + np.random.normal(0, 15), 50)
                precio = round(25 + j * 10 + k * 0.15, 2)
                filas.append([cli, sku_nombre, mes.strftime("%Y-%m"), round(unidades), precio,
                              round(inflacion, 4), round(tc, 4), canal, grupo, pais, marca, categoria, sku_id])
    return pd.DataFrame(filas, columns=REQUIRED_COLS + ["Inflacion_Mensual", "Var_TipoCambio"]
                         + ATRIBUTOS_CLIENTE + ATRIBUTOS_SKU_BB + ATRIBUTOS_SKU_ID)


def meses_futuros(ultimo_mes: str, horizonte: int):
    inicio = pd.to_datetime(ultimo_mes) + pd.DateOffset(months=1)
    return pd.date_range(inicio, periods=horizonte, freq="MS").strftime("%Y-%m").tolist()


# -------------------------------------------------------------------
# 1. Encabezado (logo + título) y carga de datos
# -------------------------------------------------------------------
with st.sidebar:
    st.header("0. Marca")
    logo_file = st.file_uploader("Logo CBM (opcional, PNG/JPG)", type=["png", "jpg", "jpeg"], key="logo")

    st.header("👥 Modo colaborativo")
    if GSHEETS_DISPONIBLE and hay_credenciales_gsheets():
        modo_colaborativo = st.checkbox(
            "Guardar y compartir cambios con todo el equipo (Google Sheets)", value=True,
            help="Al guardar, cualquier persona que abra este link después verá los últimos valores guardados.",
        )
        conn_sheets = st.connection("gsheets", type=GSheetsConnection) if modo_colaborativo else None
    else:
        modo_colaborativo = False
        conn_sheets = None
        st.caption("⚠️ Modo colaborativo no configurado — cada quien trabaja solo en su sesión. "
                   "Ver README para conectar un Google Sheet compartido.")

col_logo, col_title = st.columns([1, 6])
with col_logo:
    if logo_file is not None:
        st.image(logo_file, width=90)
with col_title:
    st.title("Modelo de Proyección de Presupuesto Colaborativo CBM - CAM SUR")
st.caption("Regresión lineal + variables macro editables + % colaborativo por iniciativa comercial")

with st.sidebar:
    st.header("1. Datos históricos")
    archivo = st.file_uploader("Sube tu histórico (CSV o Excel)", type=["csv", "xlsx"])
    usar_demo = st.checkbox("Usar datos de ejemplo (ficticios)", value=(archivo is None))

def limpiar_numero(serie: pd.Series) -> pd.Series:
    """Convierte una columna de texto a número, tolerando formatos comunes de Excel/CSV
    latinoamericano: espacios como separador de miles ('1 152'), coma como decimal ('28,50'),
    punto como separador de miles con coma decimal ('1.234,56'), o ya numérico de por sí."""
    if pd.api.types.is_numeric_dtype(serie):
        return serie
    s = serie.astype(str).str.strip()
    s = s.str.replace("\xa0", "", regex=False).str.replace(" ", "", regex=False)  # espacios de miles
    tiene_coma = s.str.contains(",", regex=False)
    tiene_punto = s.str.contains(".", regex=False)
    ambos = tiene_coma & tiene_punto
    # Si trae los dos separadores, el que aparece último es el decimal (ej. '1.234,56' -> coma es decimal)
    s_ambos_coma_decimal = s.where(~ambos, s.str.replace(".", "", regex=False))
    s = s_ambos_coma_decimal.str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce")


def cargar_historico(archivo):
    """Lee CSV o Excel de forma robusta: detecta separador de columnas (coma/punto y coma/tab) y
    codificación (utf-8/latin-1) automáticamente, limpia espacios en los nombres de columna (ej.
    'Cliente; SKU' con espacio tras el ';'), y limpia los valores numéricos de Unidades/Precio y de
    cualquier variable macro tolerando espacios de miles y coma decimal — formatos típicos de
    Excel/ERP en español."""

    def limpiar_columnas(df):
        df.columns = [str(c).strip() for c in df.columns]
        return df

    if not archivo.name.endswith(".csv"):
        df = limpiar_columnas(pd.read_excel(archivo))
    else:
        import csv
        ultimo_error, df = None, None
        for encoding in ("utf-8-sig", "utf-8", "latin-1"):
            try:
                archivo.seek(0)
                muestra = archivo.read(8192)
                if isinstance(muestra, bytes):
                    muestra = muestra.decode(encoding)
                delimitador = csv.Sniffer().sniff(muestra, delimiters=",;\t").delimiter
                archivo.seek(0)
                df = limpiar_columnas(pd.read_csv(archivo, sep=delimitador, encoding=encoding,
                                                   skipinitialspace=True, dtype=str))
                break
            except Exception as e:
                ultimo_error = e
        if df is None:
            for encoding in ("utf-8-sig", "utf-8", "latin-1"):
                try:
                    archivo.seek(0)
                    df = limpiar_columnas(pd.read_csv(archivo, sep=None, engine="python",
                                                        encoding=encoding, dtype=str))
                    break
                except Exception as e:
                    ultimo_error = e
        if df is None:
            raise ultimo_error

    # Limpieza numérica: Unidades, Precio, y cualquier otra columna que no sea texto por diseño
    # (Cliente/SKU/Mes y los atributos de referencia se dejan tal cual, son texto legítimo)
    columnas_texto = {"Cliente", "SKU", "Mes"} | set(ATRIBUTOS_CLIENTE) | set(ATRIBUTOS_SKU)
    for col in df.columns:
        if col in ("Unidades", "Precio"):
            df[col] = limpiar_numero(df[col])
        elif col not in columnas_texto:
            candidata = limpiar_numero(df[col])
            # Solo adoptamos la versión numérica si casi todo convirtió bien (si no, es texto legítimo)
            if candidata.notna().mean() >= 0.9:
                df[col] = candidata
    return df


if archivo is not None and not usar_demo:
    try:
        hist = cargar_historico(archivo)
    except Exception as e:
        st.error(
            "No pude leer tu archivo. Esto casi siempre pasa por el formato del CSV — por ejemplo, "
            "si Excel lo guardó separado por punto y coma (`;`) en vez de coma, o con una codificación "
            "de caracteres distinta a UTF-8.\n\n"
            "**Prueba esto:**\n"
            "1. En Excel: 'Guardar como' → elige **'Libro de Excel (.xlsx)'** en vez de CSV, y sube ese archivo. "
            "Es el formato más confiable.\n"
            "2. Si necesitas CSV sí o sí: ábrelo en Excel, 'Guardar como' → 'CSV UTF-8 (delimitado por comas)'.\n\n"
            f"Detalle técnico del error: `{e}`"
        )
        st.stop()
else:
    hist = datos_demo()

faltantes = [c for c in REQUIRED_COLS if c not in hist.columns]
if faltantes:
    st.error(f"Faltan columnas obligatorias en tu archivo: {faltantes}. "
             f"Se requieren al menos: {REQUIRED_COLS}")
    st.stop()

atributos_cliente_presentes = [c for c in ATRIBUTOS_CLIENTE if c in hist.columns]
atributos_sku_presentes = [c for c in ATRIBUTOS_SKU if c in hist.columns]
atributos_presentes = atributos_cliente_presentes + atributos_sku_presentes  # todas las dims de building blocks
posibles_macro = [
    c for c in hist.columns
    if c not in REQUIRED_COLS and c not in ATRIBUTOS_BB and pd.api.types.is_numeric_dtype(hist[c])
]

faltan_bb = [c for c in ATRIBUTOS_BB if c not in hist.columns]
if faltan_bb:
    st.sidebar.info(f"Tip: agrega columnas {faltan_bb} a tu histórico para habilitar todos los building "
                     f"blocks (Canal, Grupo_Cliente, Pais, Marca, Categoria). No son necesarias para el modelo de proyección.")

hist["Venta"] = (hist["Unidades"] * hist["Precio"]).round(2)

# Si el archivo trae varias filas para el mismo Cliente-SKU-Mes (ej. transacciones sueltas en vez de
# un total mensual ya agregado), las consolidamos en una sola fila por mes. Si no hay duplicados,
# esto no cambia nada.
n_antes = len(hist)
cols_agg_extra = {c: "mean" for c in posibles_macro}
cols_agg_extra.update({c: "first" for c in atributos_presentes})
hist = hist.groupby(["Cliente", "SKU", "Mes"], as_index=False).agg(
    Unidades=("Unidades", "sum"), Venta=("Venta", "sum"), **{
        k: pd.NamedAgg(column=k, aggfunc=v) for k, v in cols_agg_extra.items()
    }
)
hist["Precio"] = np.where(hist["Unidades"] != 0, hist["Venta"] / hist["Unidades"], 0.0).round(2)
if n_antes != len(hist):
    st.sidebar.info(f"Tu archivo traía {n_antes - len(hist)} filas duplicadas para el mismo Cliente-SKU-Mes "
                     f"(ej. transacciones sueltas) — las consolidé sumando Unidades y Venta por mes.")

with st.sidebar:
    st.header("2. Variables macro / adicionales")
    macro_cols = st.multiselect(
        "Selecciona las columnas que usaremos como variables explicativas del modelo",
        options=posibles_macro, default=posibles_macro,
    )
    st.header("3. Horizonte de forecast")
    horizonte = st.slider("Meses a proyectar", 1, 24, 12)

st.subheader("Histórico cargado")
st.dataframe(hist.head(20), use_container_width=True)

# -------------------------------------------------------------------
# 2. Ajuste de Regresión Lineal por Cliente-SKU
# -------------------------------------------------------------------
hist = hist.sort_values(["Cliente", "SKU", "Mes"]).reset_index(drop=True)
combos = hist[["Cliente", "SKU"]].drop_duplicates().values.tolist()
ultimo_mes_global = hist["Mes"].max()
meses_fcst = meses_futuros(ultimo_mes_global, horizonte)

cliente_attrs = hist[["Cliente"] + atributos_cliente_presentes].drop_duplicates("Cliente") if atributos_cliente_presentes else None
sku_attrs = hist[["SKU"] + atributos_sku_presentes].drop_duplicates("SKU") if atributos_sku_presentes else None


@st.cache_data(show_spinner=False)
def ajustar_modelos(hist_df: pd.DataFrame, macro_cols: list, meses_fcst: list, combos: list):
    """Enfoque top-down: se agregan TODOS los Cliente-SKU en una sola serie mensual total,
    se ajusta un único modelo de tendencia + estacionalidad (+ macro) sobre esa serie agregada
    —mucho más estable y con estacionalidad real y visible, al no diluirse en miles de series
    individuales con poca historia— y luego se desagrega el resultado por la participación
    histórica de cada Cliente-SKU dentro del total. Los de compra única (1 solo mes) quedan
    excluidos (no se proyectan)."""
    CAP_TREND_CON_MACRO = 0.03   # tope de tendencia: 3%/mes si hay variables macro explicando el crecimiento
    CAP_TREND_SIN_MACRO = 0.01   # 1%/mes si NO hay macro — sin explicación, el BAU debe ser casi plano
    CAP_ESTACIONALIDAD = 0.25    # con pocos años de historia (ej. 2), cada mes calendario tiene muy pocos
                                  # datos (~2) para estimar su estacionalidad — se topa la amplitud para que
                                  # no se sobreajuste a ruido y termine inventando un pico/valle irreal

    referencias = {col: hist_df[col].tail(6).mean() for col in macro_cols}

    # Índice calendario global (continuo, sin huecos) para que la tendencia y la estacionalidad
    # se calculen sobre el mes real, no sobre la posición de la fila.
    todos_los_meses = sorted(set(hist_df["Mes"].unique()) | set(meses_fcst))
    base_y, base_m = (int(x) for x in todos_los_meses[0].split("-"))

    def idx_calendario(mes_str):
        y, m = (int(x) for x in mes_str.split("-"))
        return (y - base_y) * 12 + (m - base_m)

    def mes_del_anio(mes_str):
        return int(mes_str.split("-")[1])

    # ---------------------------------------------------------------
    # PASO 1: un único modelo sobre la serie TOTAL agregada (todos los Cliente-SKU sumados)
    # ---------------------------------------------------------------
    total_mensual = hist_df.groupby("Mes", as_index=False)["Unidades"].sum().sort_values("Mes")
    y_hist_total = total_mensual["Unidades"].values.astype(float)
    idx_hist_total = np.array([idx_calendario(m) for m in total_mensual["Mes"]])
    mes_hist_total = np.array([mes_del_anio(m) for m in total_mensual["Mes"]])
    sin_h, cos_h = np.sin(2 * np.pi * mes_hist_total / 12), np.cos(2 * np.pi * mes_hist_total / 12)

    if macro_cols:
        macro_por_mes = hist_df.groupby("Mes")[macro_cols].mean()
        macro_hist_total = np.column_stack([macro_por_mes.reindex(total_mensual["Mes"])[c].values for c in macro_cols])
    else:
        macro_hist_total = np.empty((len(total_mensual), 0))

    X_total = np.column_stack([idx_hist_total, sin_h, cos_h, macro_hist_total])
    y_log_total = np.log1p(np.clip(y_hist_total, 0, None))
    modelo_total = LinearRegression().fit(X_total, y_log_total)

    coefs = dict(zip(["trend", "sin", "cos"] + macro_cols, modelo_total.coef_))
    coefs["intercept"] = modelo_total.intercept_
    cap = CAP_TREND_CON_MACRO if macro_cols else CAP_TREND_SIN_MACRO
    coefs["trend"] = float(np.clip(coefs["trend"], -cap, cap))

    amplitud_estacional = float(np.hypot(coefs["sin"], coefs["cos"]))
    if amplitud_estacional > CAP_ESTACIONALIDAD:
        factor_ajuste = CAP_ESTACIONALIDAD / amplitud_estacional
        coefs["sin"] *= factor_ajuste
        coefs["cos"] *= factor_ajuste

    # Al topar tendencia/estacionalidad DESPUÉS de ajustar el modelo, el intercepto original queda
    # descalibrado (fue estimado para acompañar los coeficientes sin topar). Se recalibra anclando
    # al nivel RECIENTE (últimos 6 meses, no los 2 años completos) — así el BAU parte de "dónde está
    # el negocio hoy", con una pendiente/estacionalidad conservadora encima, en vez de regresar hacia
    # el promedio histórico completo si hubo crecimiento real en el camino.
    coef_vec_hist = np.array([coefs["trend"], coefs["sin"], coefs["cos"]] + [coefs[c] for c in macro_cols])
    componente_capado = X_total @ coef_vec_hist
    residuales_recientes = (y_log_total - componente_capado)[-6:]
    coefs["intercept"] = float(np.mean(residuales_recientes))

    idx_futuro = np.array([idx_calendario(m) for m in meses_fcst])
    mes_futuro = np.array([mes_del_anio(m) for m in meses_fcst])
    sin_f, cos_f = np.sin(2 * np.pi * mes_futuro / 12), np.cos(2 * np.pi * mes_futuro / 12)
    macro_futuro = np.array([[referencias[c]] * len(meses_fcst) for c in macro_cols]).T if macro_cols else np.empty((len(meses_fcst), 0))
    X_f_total = np.column_stack([idx_futuro, sin_f, cos_f, macro_futuro])
    coef_vec = np.array([coefs["trend"], coefs["sin"], coefs["cos"]] + [coefs[c] for c in macro_cols])
    y_log_pred_total = coefs["intercept"] + X_f_total @ coef_vec
    bau_total_futuro = np.expm1(y_log_pred_total)

    # Tope de sensatez sobre el nivel total: no más de 2x el máximo histórico ni 3x el promedio.
    tope_total = max(y_hist_total.max() * 2, y_hist_total.mean() * 3, 1.0)
    bau_total_futuro = np.clip(bau_total_futuro, 0, tope_total)
    bau_total_dict = dict(zip(meses_fcst, bau_total_futuro))

    # ---------------------------------------------------------------
    # PASO 2: desagregar por participación histórica de cada Cliente-SKU dentro del total
    # ---------------------------------------------------------------
    conteo_meses = hist_df.groupby(["Cliente", "SKU"]).size()
    vol_por_combo = hist_df.groupby(["Cliente", "SKU"])["Unidades"].sum()
    vol_total_hist = hist_df["Unidades"].sum()

    resultados_bau, coeficientes = [], []
    for cli, sku in combos:
        n = conteo_meses.get((cli, sku), 0)
        vol = vol_por_combo.get((cli, sku), 0.0)
        if n <= 1 or vol_total_hist <= 0:
            participacion = 0.0
            categoria = "Excluido (compra única)"
        else:
            participacion = vol / vol_total_hist
            categoria = "Participación histórica (modelo top-down)"

        coeficientes.append({"Cliente": cli, "SKU": sku, "Participacion_%": round(participacion * 100, 4), **coefs})
        for mes in meses_fcst:
            resultados_bau.append({"Cliente": cli, "SKU": sku, "Mes": mes,
                                    "BAU_Unidades": round(participacion * bau_total_dict[mes]),
                                    "Categoria_Modelo": categoria})

    return pd.DataFrame(resultados_bau), pd.DataFrame(coeficientes), referencias


bau_df, coef_df, referencias = ajustar_modelos(hist, macro_cols, meses_fcst, combos)

st.subheader("Modelo ajustado (top-down: un único modelo sobre el total, desagregado por participación)")
coefs_unicos = coef_df.drop(columns=["Cliente", "SKU", "Participacion_%"]).drop_duplicates().round(4)
st.dataframe(coefs_unicos, use_container_width=True)
st.caption("log(Unidades totales) = intercepto + tendencia·mes_calendario + sin/cos(mes_del_año) + "
           "Σ(coef_macro × variable_macro). Un solo modelo para TODOS los Cliente-SKU juntos — más estable "
           "que ajustar miles de modelos individuales con poca historia cada uno.")

with st.expander("Ver participación histórica de cada Cliente-SKU (cómo se reparte el total)"):
    st.dataframe(coef_df[["Cliente", "SKU", "Participacion_%"]].sort_values("Participacion_%", ascending=False),
                 use_container_width=True)

# -------------------------------------------------------------------
# 3. Escenario macro editable
# -------------------------------------------------------------------
st.subheader("4. Escenario macro futuro (editable)")
if macro_cols:
    escenario_default = pd.DataFrame({"Mes": meses_fcst})
    for c in macro_cols:
        escenario_default[c] = referencias[c]
    if modo_colaborativo:
        escenario_default = leer_tabla_compartida(conn_sheets, "escenario_macro", escenario_default,
                                                   claves=["Mes"], nombre_legible="Escenario macro")
    escenario = st.data_editor(escenario_default, num_rows="fixed", use_container_width=True,
                                key="escenario_macro")
else:
    escenario = pd.DataFrame({"Mes": meses_fcst})
    st.info("No seleccionaste variables macro — el forecast usará solo tendencia y estacionalidad.")

# -------------------------------------------------------------------
# 4. Palancas de Ejecución Propia — editable por Cliente-SKU-Mes
#    (mismo lenguaje que el puente de crecimiento del presupuesto: Nuevos Productos,
#    Profundización de Cuentas, Cuentas Nuevas)
# -------------------------------------------------------------------
st.subheader("5. Palancas de Ejecución Propia (editable)")
st.caption("Equivalente a las palancas 3, 4 y 5 del puente de crecimiento del presupuesto. "
           "Afectan **volumen** (unidades), no precio.")
PALANCAS_EJECUCION = ["Nuevos_Productos_%", "Profundizacion_Cuentas_%", "Cuentas_Nuevas_%"]
inic_default = bau_df[["Cliente", "SKU", "Mes"]].copy()
for pal in PALANCAS_EJECUCION:
    inic_default[pal] = 0.0
inic_default["Iniciativa"] = ""
if modo_colaborativo:
    inic_default = leer_tabla_compartida(conn_sheets, "iniciativas_colaborativo", inic_default,
                                          claves=["Cliente", "SKU", "Mes"], nombre_legible="Palancas de ejecución")
iniciativas = st.data_editor(
    inic_default, num_rows="fixed", use_container_width=True, key="iniciativas",
    column_config={pal: st.column_config.NumberColumn(label=pal.replace("_", " "), format="%.1f%%", step=0.5)
                   for pal in PALANCAS_EJECUCION},
)

# -------------------------------------------------------------------
# 5. Precio: híbrido = escalación automática por inflación + ajuste manual opcional
# -------------------------------------------------------------------
st.subheader("6. Precio proyectado (híbrido: automático + ajuste manual)")

col_a, col_b = st.columns([1, 1])
with col_a:
    st.markdown("**6a. Precio base por SKU (editable)**")

    def _precio_ponderado_reciente(g):
        g = g.sort_values("Mes").tail(6)
        u = g["Unidades"].sum()
        return (g["Unidades"] * g["Precio"]).sum() / u if u != 0 else g["Precio"].tail(6).median()

    precio_default = (
        hist.groupby("SKU")[["Mes", "Unidades", "Precio"]]
        .apply(_precio_ponderado_reciente, include_groups=False)
        .reset_index().rename(columns={0: "Precio_Base"})
    )
    if modo_colaborativo:
        precio_default = leer_tabla_compartida(conn_sheets, "precios_base", precio_default,
                                                claves=["SKU"], nombre_legible="Precio base")
    precios_base = st.data_editor(precio_default, num_rows="fixed", use_container_width=True, key="precios_base")
    st.caption("Se usa el precio ponderado por volumen real (Venta ÷ Unidades) de los últimos 6 meses por SKU "
               "— no el último dato puntual — para que una transacción atípica o un renglón que no es un "
               "precio unitario real (ej. un monto de descuento o servicio) no distorsione todo el año proyectado.")

with col_b:
    st.markdown("**6b. Tasa de inflación mensual para escalar el precio**")
    tasa_sugerida = float(referencias.get("Inflacion_Mensual", 0.0) * 100) if "Inflacion_Mensual" in referencias else 0.5
    tasa_inflacion_precio = st.number_input(
        "Tasa mensual (%)", min_value=-5.0, max_value=10.0, value=round(tasa_sugerida, 2), step=0.05,
        help="El precio base se escala mes a mes con esta tasa compuesta. "
             "Por defecto toma el promedio histórico de tu variable de inflación, si la seleccionaste como macro.",
    )
    st.caption(f"Precio_Mes_n = Precio_Base × (1 + {tasa_inflacion_precio:.2f}%)ⁿ")

st.markdown("**6c. Ajuste manual de precio por Cliente-SKU-Mes (editable, 0% = usa solo el escalado automático)**")
st.caption("Úsalo solo para casos puntuales: aumento de lista negociado con un cliente específico, "
           "cambio de precio por contrato, etc. En el 95% de los casos lo dejas en 0%.")
ajuste_precio_default = bau_df[["Cliente", "SKU", "Mes"]].copy()
ajuste_precio_default["Ajuste_Manual_Precio_%"] = 0.0
ajuste_precio_default["Motivo"] = ""
if modo_colaborativo:
    ajuste_precio_default = leer_tabla_compartida(conn_sheets, "ajuste_manual_precio", ajuste_precio_default,
                                                   claves=["Cliente", "SKU", "Mes"], nombre_legible="Ajuste manual de precio")
ajuste_precio = st.data_editor(
    ajuste_precio_default, num_rows="fixed", use_container_width=True, key="ajuste_precio",
    column_config={"Ajuste_Manual_Precio_%": st.column_config.NumberColumn(format="%.1f%%", step=0.5)},
)

if modo_colaborativo:
    st.markdown("---")
    c1, c2 = st.columns([1, 3])
    with c1:
        guardar = st.button("💾 Guardar cambios para todo el equipo", type="primary", use_container_width=True)
    with c2:
        st.caption("Guarda el escenario macro, el % colaborativo, el precio base y el ajuste manual "
                   "en el Google Sheet compartido. La próxima persona que abra este link verá estos valores. "
                   "Si dos personas guardan al mismo tiempo, gana la última en guardar. Para eficiencia, "
                   "solo se guardan las filas que sí editaste (no las miles que quedan en 0%).")
    if guardar:
        with st.spinner("Guardando..."):
            guardar_tabla_compartida(conn_sheets, "escenario_macro", escenario)

            iniciativas_editadas = iniciativas[
                (iniciativas[PALANCAS_EJECUCION] != 0).any(axis=1) | (iniciativas["Iniciativa"].astype(str).str.strip() != "")
            ]
            guardar_tabla_compartida(conn_sheets, "iniciativas_colaborativo", iniciativas_editadas)

            guardar_tabla_compartida(conn_sheets, "precios_base", precios_base)

            ajuste_precio_editado = ajuste_precio[
                (ajuste_precio["Ajuste_Manual_Precio_%"] != 0) | (ajuste_precio["Motivo"].astype(str).str.strip() != "")
            ]
            guardar_tabla_compartida(conn_sheets, "ajuste_manual_precio", ajuste_precio_editado)
        n_editadas = len(iniciativas_editadas) + len(ajuste_precio_editado)
        st.success(f"✅ Guardado {datetime.now():%Y-%m-%d %H:%M} — {n_editadas} filas editadas guardadas, "
                   f"ya visibles para todo el equipo.")

# -------------------------------------------------------------------
# 6. Cálculo del forecast final
# -------------------------------------------------------------------
# Renombramos los coeficientes del modelo para que no choquen con los nombres
# de las variables macro (el escenario usa el nombre original, ej. "Inflacion_Mensual";
# el coeficiente del modelo pasa a llamarse "coef_Inflacion_Mensual").
coef_df_ren = coef_df.rename(columns={c: f"coef_{c}" for c in macro_cols})

final = bau_df.merge(escenario, on="Mes", how="left") if macro_cols else bau_df.copy()
final = final.merge(coef_df_ren, on=["Cliente", "SKU"], how="left")
final = final.merge(iniciativas[["Cliente", "SKU", "Mes"] + PALANCAS_EJECUCION], on=["Cliente", "SKU", "Mes"], how="left")
final = final.merge(precios_base, on="SKU", how="left")
final = final.merge(ajuste_precio[["Cliente", "SKU", "Mes", "Ajuste_Manual_Precio_%"]], on=["Cliente", "SKU", "Mes"], how="left")

# Red de seguridad: si algún merge no encontró coincidencia (por ejemplo, datos de una corrida
# anterior con otro archivo que no se limpiaron a tiempo), rellenamos con un valor neutro en vez
# de dejar NaN — así nunca vuelve a tronar el cálculo final, aunque sí conviene revisar el aviso
# de la barra lateral si aparece.
for c in macro_cols:
    final[c] = final[c].fillna(referencias[c])
    final[f"coef_{c}"] = final[f"coef_{c}"].fillna(0.0)
final["BAU_Unidades"] = final["BAU_Unidades"].fillna(0)
if final["Precio_Base"].isna().any():
    st.sidebar.warning("⚠️ Algunos SKU no tenían Precio_Base guardado — se usó el precio promedio como respaldo.")
    final["Precio_Base"] = final["Precio_Base"].fillna(final["Precio_Base"].mean())

# Número de mes transcurrido (1, 2, 3...) para la escalación compuesta
mes_orden = {m: i + 1 for i, m in enumerate(meses_fcst)}
final["n_mes"] = final["Mes"].map(mes_orden)
final["Precio_Escalado"] = final["Precio_Base"] * (1 + tasa_inflacion_precio / 100.0) ** final["n_mes"]
final["Precio"] = (final["Precio_Escalado"] * (1 + final["Ajuste_Manual_Precio_%"].fillna(0) / 100.0)).round(2)

# Palanca "Mercado Orgánico" = Σ coef_i × (valor_escenario_i − valor_referencia_i), en % (el modelo
# ahora corre en espacio log, así que los coeficientes de macro ya representan %-efecto, no unidades).
final["Mercado_Organico_%"] = 0.0
for c in macro_cols:
    final["Mercado_Organico_%"] += (final[c] - referencias[c]) * final[f"coef_{c}"]

for pal in PALANCAS_EJECUCION:
    final[pal] = final[pal].fillna(0)
final["Ejecucion_Propia_%"] = final[PALANCAS_EJECUCION].sum(axis=1)

final["Unidades_Final"] = np.round(
    (final["BAU_Unidades"] * (1 + final["Mercado_Organico_%"])).clip(lower=0)
    * (1 + final["Ejecucion_Propia_%"] / 100.0)
).astype(int)
final["Venta_Final"] = (final["Unidades_Final"] * final["Precio"]).round(0)

if cliente_attrs is not None:
    final = final.merge(cliente_attrs, on="Cliente", how="left")
if sku_attrs is not None:
    final = final.merge(sku_attrs, on="SKU", how="left")

st.subheader("7. Forecast Final")
atributos_sku_id_presentes = [c for c in ATRIBUTOS_SKU_ID if c in hist.columns]
cols_contexto = [c for c in ["SKU_ID"] + atributos_cliente_presentes + [x for x in ATRIBUTOS_SKU_BB if x in atributos_sku_presentes]
                 if c in final.columns]
cols_mostrar = (["Cliente", "SKU"] + cols_contexto +
                ["Mes", "Categoria_Modelo", "Participacion_%", "BAU_Unidades", "Mercado_Organico_%"] + PALANCAS_EJECUCION +
                ["Unidades_Final", "Precio_Base", "Ajuste_Manual_Precio_%", "Precio", "Venta_Final"])
st.dataframe(final[cols_mostrar], use_container_width=True)
st.caption("El modelo es 'top-down': se ajusta tendencia + estacionalidad una sola vez sobre el TOTAL de todos "
           "los Cliente-SKU juntos (mucho más estable que miles de series individuales con poca historia), y "
           "ese total se reparte según la Participacion_% histórica de cada Cliente-SKU. Categoria_Modelo indica "
           "si esa combinación se 'Excluyó' (solo 1 mes de historia, no se proyecta) o si tiene 'Participación "
           "histórica' (sí se proyecta, según su peso dentro del total). Precio = Precio_Base escalado por la "
           "tasa de inflación mensual (6b) × (1 + Ajuste_Manual_Precio_% del mes). Mercado_Organico_% = palanca "
           "'Mercado Orgánico' del puente de crecimiento (efecto de tus variables macro, en %, aplicado sobre el "
           "total). Pais, Canal, Grupo_Cliente, Marca, Categoria y SKU_ID son características de referencia — no "
           "entran al modelo de proyección.")

with st.expander("¿Cuántos Cliente-SKU cayeron en cada categoría del modelo?"):
    resumen_categoria = final.drop_duplicates(["Cliente", "SKU"])["Categoria_Modelo"].value_counts().reset_index()
    resumen_categoria.columns = ["Categoría", "Cantidad de Cliente-SKU"]
    st.dataframe(resumen_categoria, use_container_width=True)

resumen = final.groupby("Cliente").agg(
    Unidades_BAU=("BAU_Unidades", "sum"),
    Unidades_Final=("Unidades_Final", "sum"),
    Venta_Final=("Venta_Final", "sum"),
).reset_index()
resumen["Impacto_%"] = (resumen["Unidades_Final"] - resumen["Unidades_BAU"]) / resumen["Unidades_BAU"]

st.subheader("8. Resumen Ejecutivo por Cliente")
st.dataframe(resumen, use_container_width=True)
st.bar_chart(resumen.set_index("Cliente")[["Unidades_BAU", "Unidades_Final"]])

# -------------------------------------------------------------------
# 6.5 Construir dataset combinado Real + Proyectado (para tendencia y building blocks)
# -------------------------------------------------------------------
hist_comb = hist[["Cliente", "SKU", "Mes", "Unidades", "Venta"] + atributos_presentes].copy()
hist_comb["Tipo"] = "Real"

fcst_comb_cols = ["Cliente", "SKU", "Mes", "Unidades_Final", "Venta_Final"] + atributos_presentes
fcst_comb = final[fcst_comb_cols].rename(columns={"Unidades_Final": "Unidades", "Venta_Final": "Venta"}).copy()
fcst_comb["Tipo"] = "Proyectado"

combinado = pd.concat([hist_comb, fcst_comb], ignore_index=True)
orden_meses = sorted(hist["Mes"].unique()) + list(meses_fcst)
combinado["Mes"] = pd.Categorical(combinado["Mes"], categories=orden_meses, ordered=True)

# -------------------------------------------------------------------
# 7. Tendencia de Volumen y Venta — Real + Proyectado, todos los periodos
# -------------------------------------------------------------------
st.subheader("9. Tendencia de Volumen y Venta — Real + Proyectado (todos los periodos)")

tendencia = combinado.groupby(["Mes", "Tipo"], observed=True).agg(
    Unidades=("Unidades", "sum"), Venta=("Venta", "sum")
).reset_index()

col_v, col_p = st.columns(2)
with col_v:
    chart_vol = alt.Chart(tendencia).mark_line(point=True).encode(
        x=alt.X("Mes:N", sort=orden_meses, title="Mes"),
        y=alt.Y("Unidades:Q", title="Unidades"),
        color=alt.Color("Tipo:N", scale=alt.Scale(domain=["Real", "Proyectado"], range=["#1F4E78", "#F28E2B"])),
        strokeDash=alt.StrokeDash("Tipo:N", scale=alt.Scale(domain=["Real", "Proyectado"], range=[[1, 0], [6, 3]])),
        tooltip=["Mes", "Tipo", "Unidades"],
    ).properties(title="Volumen (Unidades) por mes", height=350)
    st.altair_chart(chart_vol, use_container_width=True)
with col_p:
    chart_venta = alt.Chart(tendencia).mark_line(point=True).encode(
        x=alt.X("Mes:N", sort=orden_meses, title="Mes"),
        y=alt.Y("Venta:Q", title="Venta ($)"),
        color=alt.Color("Tipo:N", scale=alt.Scale(domain=["Real", "Proyectado"], range=["#1F4E78", "#F28E2B"])),
        strokeDash=alt.StrokeDash("Tipo:N", scale=alt.Scale(domain=["Real", "Proyectado"], range=[[1, 0], [6, 3]])),
        tooltip=["Mes", "Tipo", "Venta"],
    ).properties(title="Venta ($) por mes", height=350)
    st.altair_chart(chart_venta, use_container_width=True)
st.caption("Línea sólida = Real (histórico). Línea punteada = Proyectado (forecast).")

# -------------------------------------------------------------------
# 8. Building Blocks por Canal, Grupo de Clientes, Marca y Categoría
# -------------------------------------------------------------------
st.subheader("10. Building Blocks por Canal, Grupo de Clientes, Marca y Categoría")

if not atributos_presentes:
    st.warning("Tu histórico no trae ninguna de las columnas 'Canal', 'Grupo_Cliente', 'Pais', 'Marca' o 'Categoria', "
               "así que no puedo armar este desglose. Agrégalas a tu archivo y vuelve a cargarlo — Canal y "
               "Grupo_Cliente van una por Cliente, Marca y Categoria van una por SKU. No afectan el modelo de "
               "proyección, solo habilitan esta vista.")
else:
    faltan_dims = [c for c in ATRIBUTOS_BB if c not in atributos_presentes]
    if faltan_dims:
        st.caption(f"No tienes cargadas estas dimensiones, así que no se muestran: {faltan_dims}.")

    vista = st.radio(
        "Periodo para el desglose",
        ["Acumulado (Real + Proyectado)", "Solo Real (histórico)", "Solo Proyectado (forecast)", "Mes específico"],
        horizontal=True,
    )
    if vista == "Mes específico":
        mes_sel = st.selectbox("Selecciona el mes", orden_meses)
        datos_bb = combinado[combinado["Mes"] == mes_sel]
    elif vista == "Solo Real (histórico)":
        datos_bb = combinado[combinado["Tipo"] == "Real"]
    elif vista == "Solo Proyectado (forecast)":
        datos_bb = combinado[combinado["Tipo"] == "Proyectado"]
    else:
        datos_bb = combinado

    metrica_bb = st.radio("Métrica", ["Unidades", "Venta"], horizontal=True, key="metrica_bb")

    COLORES_BB = {"Canal": "#1F4E78", "Grupo_Cliente": "#F28E2B", "Pais": "#2E7D32",
                  "Marca": "#4C9A2A", "Categoria": "#A02B93"}
    TITULOS_BB = {"Canal": "Canal", "Grupo_Cliente": "Grupo de Clientes", "Pais": "País",
                  "Marca": "Marca", "Categoria": "Categoría"}

    dims_a_mostrar = [d for d in ATRIBUTOS_BB if d in atributos_presentes]
    for fila_inicio in range(0, len(dims_a_mostrar), 2):
        cols = st.columns(2)
        for idx, dim in enumerate(dims_a_mostrar[fila_inicio:fila_inicio + 2]):
            with cols[idx]:
                bb = (datos_bb.groupby(dim, observed=True)[metrica_bb].sum()
                      .reset_index().sort_values(metrica_bb, ascending=False))
                chart_bb = alt.Chart(bb).mark_bar(color=COLORES_BB[dim]).encode(
                    x=alt.X(f"{dim}:N", sort="-y", title=TITULOS_BB[dim]),
                    y=alt.Y(f"{metrica_bb}:Q"),
                    tooltip=[dim, metrica_bb],
                ).properties(title=f"{metrica_bb} por {TITULOS_BB[dim]}", height=330)
                st.altair_chart(chart_bb, use_container_width=True)

# -------------------------------------------------------------------
# 9. Puente de Crecimiento — mismo lenguaje que la descomposición del presupuesto
#    (Base -> Precio -> Mercado Orgánico -> Nuevos Productos -> Profundización -> Cuentas Nuevas -> Total)
# -------------------------------------------------------------------
st.subheader("11. Puente de Crecimiento (Venta, todo el horizonte proyectado)")
st.caption("Misma estructura que la descomposición del presupuesto: parte de una base sin ajustes y va "
           "sumando cada palanca en orden, hasta llegar a la Venta Final proyectada.")

venta_base = float((final["BAU_Unidades"] * final["Precio_Base"]).sum())
unidades_tras_mercado = (final["BAU_Unidades"] * (1 + final["Mercado_Organico_%"])).clip(lower=0)
venta_tras_precio = float((final["BAU_Unidades"] * final["Precio"]).sum())
venta_tras_mercado = float((unidades_tras_mercado * final["Precio"]).sum())

unidades_cursor = unidades_tras_mercado.copy()
ventas_por_palanca = {}
venta_cursor = venta_tras_mercado
for pal in PALANCAS_EJECUCION:
    unidades_cursor = unidades_cursor * (1 + final[pal] / 100.0)
    venta_nueva = float((unidades_cursor * final["Precio"]).sum())
    ventas_por_palanca[pal] = venta_nueva - venta_cursor
    venta_cursor = venta_nueva
venta_final_total = venta_cursor

NOMBRES_PALANCA = {"Nuevos_Productos_%": "Nuevos Productos", "Profundizacion_Cuentas_%": "Profundización Cuentas",
                    "Cuentas_Nuevas_%": "Cuentas Nuevas"}

pasos = ["Base (BAU)", "Precio", "Mercado Orgánico"] + [NOMBRES_PALANCA[p] for p in PALANCAS_EJECUCION] + ["Total Proyectado"]
deltas = [venta_base, venta_tras_precio - venta_base, venta_tras_mercado - venta_tras_precio] + \
         [ventas_por_palanca[p] for p in PALANCAS_EJECUCION] + [None]
tipos = ["base"] + ["incremento"] * (len(pasos) - 2) + ["total"]

cumulativo, bottoms, tops = 0.0, [], []
for i, (paso, delta) in enumerate(zip(pasos, deltas)):
    if tipos[i] == "base":
        bottoms.append(0.0); tops.append(delta); cumulativo = delta
    elif tipos[i] == "total":
        bottoms.append(0.0); tops.append(cumulativo)
    else:
        bottoms.append(cumulativo); tops.append(cumulativo + delta); cumulativo += delta

puente_df = pd.DataFrame({"Paso": pasos, "Bottom": bottoms, "Top": tops, "Tipo": tipos})
puente_df["Valor"] = [venta_base, venta_tras_precio - venta_base, venta_tras_mercado - venta_tras_precio] + \
                      [ventas_por_palanca[p] for p in PALANCAS_EJECUCION] + [venta_final_total]
puente_df["Pct_vs_Base"] = puente_df["Valor"] / venta_base * 100 if venta_base else 0

chart_puente = alt.Chart(puente_df).mark_bar().encode(
    x=alt.X("Paso:N", sort=pasos, title=None),
    y=alt.Y("Bottom:Q", title="Venta ($)"),
    y2="Top:Q",
    color=alt.Color("Tipo:N", scale=alt.Scale(domain=["base", "incremento", "total"],
                                               range=["#1F4E78", "#F28E2B", "#2E7D32"]), legend=None),
    tooltip=["Paso", alt.Tooltip("Valor:Q", format=",.0f"), alt.Tooltip("Pct_vs_Base:Q", format=".1f")],
).properties(height=380)
st.altair_chart(chart_puente, use_container_width=True)

tabla_puente = puente_df[["Paso", "Valor", "Pct_vs_Base"]].rename(
    columns={"Valor": "Venta ($)", "Pct_vs_Base": "% vs Base"})
st.dataframe(tabla_puente.style.format({"Venta ($)": "{:,.0f}", "% vs Base": "{:.1f}%"}), use_container_width=True)
st.caption("El Total Proyectado de este puente puede diferir en centavos de la Venta Final del Resumen Ejecutivo: "
           "el puente usa unidades continuas paso a paso, mientras que la tabla de Forecast Final redondea "
           "unidades a números enteros mes a mes. La diferencia es de redondeo, no de metodología.")

# -------------------------------------------------------------------
# 10. Exportar a Excel
# -------------------------------------------------------------------
def exportar_excel(hist, coef_df, escenario, iniciativas, precios_base, ajuste_precio, tasa_inflacion_precio,
                    final, resumen, puente_df) -> bytes:
    wb = Workbook()
    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", start_color="1F4E78")

    def volcar(ws, df, start_row=1):
        for c, col in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=c, value=col)
            cell.font = hdr_font
            cell.fill = hdr_fill
        for i, row in df.iterrows():
            for c, col in enumerate(df.columns, 1):
                ws.cell(row=start_row + 1 + i, column=c, value=row[col])
        for c in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(c)].width = 20

    ws = wb.active
    ws.title = "Historico"
    volcar(ws, hist)

    ws = wb.create_sheet("Modelo_Coeficientes")
    volcar(ws, coef_df.round(4))

    ws = wb.create_sheet("Escenario_Macro")
    volcar(ws, escenario)

    ws = wb.create_sheet("Iniciativas_Colaborativo")
    volcar(ws, iniciativas)

    ws = wb.create_sheet("Precio_Base")
    ws.cell(row=1, column=1, value=f"Tasa de inflación mensual usada para escalar precio: {tasa_inflacion_precio:.2f}%").font = Font(bold=True)
    volcar(ws, precios_base, start_row=3)

    ws = wb.create_sheet("Ajuste_Manual_Precio")
    volcar(ws, ajuste_precio)

    ws = wb.create_sheet("Forecast_Final")
    volcar(ws, final[cols_mostrar])

    ws = wb.create_sheet("Resumen_Ejecutivo")
    volcar(ws, resumen.round(3))

    ws = wb.create_sheet("Puente_Crecimiento")
    volcar(ws, puente_df[["Paso", "Valor", "Pct_vs_Base"]].round(2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


excel_bytes = exportar_excel(hist, coef_df, escenario, iniciativas, precios_base, ajuste_precio,
                              tasa_inflacion_precio, final, resumen, puente_df)
csv_bytes = final[cols_mostrar].to_csv(index=False, sep=";", decimal=",").encode("utf-8-sig")

col_dl1, col_dl2 = st.columns(2)
with col_dl1:
    st.download_button(
        "⬇️ Descargar Excel del modelo (todas las hojas)",
        data=excel_bytes,
        file_name=f"Forecast_Colaborativo_{datetime.now():%Y%m%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
with col_dl2:
    st.download_button(
        "⬇️ Descargar Forecast Final en CSV",
        data=csv_bytes,
        file_name=f"Forecast_Final_{datetime.now():%Y%m%d}.csv",
        mime="text/csv",
        use_container_width=True,
    )
st.caption("El CSV usa ';' como separador y ',' como decimal — se abre directo en Excel en español sin pasos extra.")
